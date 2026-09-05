"""Exports — PDF for anything you would forward, Excel for anything financial.

Every row and every claim carries its source reference, so a forwarded file is
still checkable: the PDF ends with a numbered reference list, the spreadsheet
has a Source column with the document name and sha prefix.
"""
from __future__ import annotations

import io
import re
from datetime import datetime, timezone
from typing import Any, Dict, List, Optional

from mangotree.config.registry import PROPERTY_INDEX
from mangotree.storage.mongo import Mongo


def _pl(pid: Optional[str]) -> str:
    if not pid:
        return "Portfolio"
    p = PROPERTY_INDEX.get(pid)
    return p.canonical_address if p else pid


# =============================================================================
# Excel
# =============================================================================

def _wb():
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    wb = Workbook()
    return wb, Font, PatternFill, Alignment


def _style_header(ws, Font, PatternFill, Alignment):
    for c in ws[1]:
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="1F6F5F")
        c.alignment = Alignment(vertical="center")
    ws.freeze_panes = "A2"
    for col in ws.columns:
        width = max(10, min(60, max(len(str(c.value or "")) for c in col) + 2))
        ws.column_dimensions[col[0].column_letter].width = width


def money_xlsx(mongo: Mongo, property_id: Optional[str]) -> bytes:
    wb, Font, PatternFill, Alignment = _wb()
    ws = wb.active
    ws.title = "Money events"
    ws.append(["Date", "Property", "Type", "Description", "Amount (USD)", "Direction", "Source document", "Source sha", "Quote"])
    q: Dict[str, Any] = {"amount": {"$type": "number"}}
    if property_id:
        q["property_id"] = property_id
    total_out = total_back = 0.0
    for e in mongo.db["timeline_events"].find(q).sort("occurred_at", 1):
        d = e.get("occurred_at")
        direction = "back" if e.get("event_type") in ("payment", "payoff") else "out" if e.get("event_type") in ("funding", "construction") else "other"
        amt = float(e["amount"])
        if direction == "out":
            total_out += amt
        elif direction == "back":
            total_back += amt
        ws.append([d.strftime("%Y-%m-%d") if hasattr(d, "strftime") else "", _pl(e.get("property_id")), e.get("event_type"), e.get("title"),
                   round(amt, 2), direction, e.get("source_name"), (e.get("source_sha") or "")[:12], (e.get("quote") or "")[:300]])
    for row in ws.iter_rows(min_row=2, min_col=5, max_col=5):
        row[0].number_format = '"$"#,##0.00'
    _style_header(ws, Font, PatternFill, Alignment)
    s = wb.create_sheet("Summary")
    s.append(["Scope", _pl(property_id)])
    s.append(["Generated", datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")])
    s.append(["Money out (funding, construction)", round(total_out, 2)])
    s.append(["Money back (payments, payoffs)", round(total_back, 2)])
    s.append(["Net", round(total_back - total_out, 2)])
    s.append(["Basis", "Quote-verified timeline events carrying an amount. Each row links to its source document by sha."])
    for r in (3, 4, 5):
        s.cell(row=r, column=2).number_format = '"$"#,##0.00'
    s.column_dimensions["A"].width = 38
    s.column_dimensions["B"].width = 60
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def tasks_xlsx(mongo: Mongo, *, property_id: Optional[str] = None, owner: Optional[str] = None) -> bytes:
    wb, Font, PatternFill, Alignment = _wb()
    ws = wb.active
    ws.title = "Tasks"
    ws.append(["Status", "Owner", "Property", "Task", "Priority", "Due", "Why", "Source", "Evidence quote", "Created", "Done by", "Done at"])
    q: Dict[str, Any] = {}
    if property_id:
        q["property_id"] = property_id
    if owner:
        q["owner"] = owner
    for t in mongo.db["tasks"].find(q).sort([("status", 1), ("due", 1)]):
        ev = (t.get("evidence") or [{}])[0]
        ws.append([t.get("status"), t.get("owner"), _pl(t.get("property_id")), t.get("title"), t.get("priority"),
                   t["due"].strftime("%Y-%m-%d") if hasattr(t.get("due"), "strftime") else "", t.get("why"), t.get("source"),
                   (ev.get("quote") or "")[:300], t["created_at"].strftime("%Y-%m-%d") if hasattr(t.get("created_at"), "strftime") else "",
                   t.get("done_by") or "", t["done_at"].strftime("%Y-%m-%d") if hasattr(t.get("done_at"), "strftime") else ""])
    _style_header(ws, Font, PatternFill, Alignment)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def portfolio_xlsx(portfolio: List[dict]) -> bytes:
    wb, Font, PatternFill, Alignment = _wb()
    ws = wb.active
    ws.title = "Portfolio"
    ws.append(["Property", "Deal type", "Health", "Health reasons", "Day count", "Documents", "Events", "Invested (ledger)", "Owed (latest balance)", "Owed as of", "Open tasks", "Suggested", "Wes done", "Wes total", "Last activity"])
    # Ledger figures; a blank cell means the documents do not establish it. Never 0.
    for p in portfolio:
        m = p.get("money") or {}
        ws.append([p["address"], p.get("deal_type"), p["health"]["level"], "; ".join(p["health"]["reasons"]), p.get("day_count"), p["documents"]["total"], p["events"],
                   round(m["invested"], 2) if isinstance(m.get("invested"), (int, float)) else "not established",
                   round(m["owed"], 2) if isinstance(m.get("owed"), (int, float)) else "not established",
                   str(m.get("owed_as_of") or "")[:10], p["tasks"]["open"], p["tasks"]["suggested"], p["wes"]["done"], p["wes"]["total"],
                   (p.get("last_activity") or "")[:10]])
    for row in ws.iter_rows(min_row=2, min_col=8, max_col=10):
        for c in row:
            c.number_format = '"$"#,##0'
    _style_header(ws, Font, PatternFill, Alignment)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# =============================================================================
# PDF
# =============================================================================

def _pdf_doc(title: str, subtitle: str):
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import mm
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, leftMargin=18 * mm, rightMargin=18 * mm, topMargin=16 * mm, bottomMargin=16 * mm, title=title)
    ss = getSampleStyleSheet()
    styles = {
        "h1": ParagraphStyle("h1", parent=ss["Title"], fontSize=17, leading=21, alignment=0, spaceAfter=2, textColor=colors.HexColor("#16161a")),
        "sub": ParagraphStyle("sub", parent=ss["Normal"], fontSize=9, textColor=colors.HexColor("#6b6b76"), spaceAfter=10),
        "h2": ParagraphStyle("h2", parent=ss["Heading2"], fontSize=11.5, leading=14, spaceBefore=10, spaceAfter=4, textColor=colors.HexColor("#1f6f5f")),
        "p": ParagraphStyle("p", parent=ss["Normal"], fontSize=10, leading=14),
        "small": ParagraphStyle("small", parent=ss["Normal"], fontSize=8, leading=10.5, textColor=colors.HexColor("#6b6b76")),
        "ref": ParagraphStyle("ref", parent=ss["Normal"], fontSize=8, leading=10.5),
    }
    story = [Paragraph(_esc(title), styles["h1"]), Paragraph(_esc(subtitle), styles["sub"])]
    return doc, story, styles, buf, Paragraph, Spacer, colors


def _esc(s: Any) -> str:
    return (str(s or "").replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;"))


_URG_COLOR = {"critical": "#c2410c", "high": "#b45309", "normal": "#1f2937", "info": "#475569", "good": "#15803d"}


def _cite_to_sup(text: str) -> str:
    return re.sub(r"\[#(\d+)\]", r"<super><font color='#1f6f5f'>\1</font></super>", _esc(text))


def answer_pdf(answer: Dict[str, Any], *, question: str, scope: str, saved_by: str = "") -> bytes:
    doc, story, st, buf, P, Spacer, colors = _pdf_doc(
        "MangoTree — Answer", f"{scope} · {datetime.now(timezone.utc):%B %d, %Y} · {('saved by ' + saved_by) if saved_by else ''}")
    story.append(P(f"<b>Question.</b> {_esc(question)}", st["p"]))
    story.append(Spacer(1, 6))
    story.append(P(_cite_to_sup(answer.get("headline", "")), st["h2"]))
    for pt in answer.get("points", []):
        color = _URG_COLOR.get(pt.get("urgency"), "#1f2937")
        story.append(P(f"<font color='{color}'>■</font> {_cite_to_sup(pt.get('text', ''))} "
                       + "".join(f"<super><font color='#1f6f5f'>{s}</font></super>" for s in pt.get("sources", []) if f"[#{s}]" not in pt.get("text", "")), st["p"]))
    # Draft answers carry the email/message text; composed answers (list, figure,
    # explain) carry a written block. Both would otherwise be lost from the PDF.
    if answer.get("draft"):
        story.append(P("Draft", st["h2"]))
        for para in str(answer["draft"]).split("\n\n"):
            if para.strip():
                story.append(P(_cite_to_sup(para.strip()).replace("\n", "<br/>"), st["p"]))
    if answer.get("composed"):
        for para in str(answer["composed"]).split("\n\n"):
            if para.strip():
                story.append(P(_cite_to_sup(para.strip()).replace("\n", "<br/>"), st["p"]))
    if answer.get("disagreements"):
        story.append(P("Where the records disagree", st["h2"]))
        for d in answer["disagreements"]:
            story.append(P("• " + _cite_to_sup(d), st["p"]))
    if answer.get("risks"):
        story.append(P("Open risks", st["h2"]))
        for r in answer["risks"]:
            story.append(P("• " + _cite_to_sup(r), st["p"]))
    if answer.get("next_actions"):
        story.append(P("Next steps", st["h2"]))
        for a in answer["next_actions"]:
            story.append(P(f"• <b>{_esc(a.get('owner'))}</b>: {_cite_to_sup(a.get('title', ''))}" + (f" (by {_esc(a['due'])})" if a.get("due") else ""), st["p"]))
    if answer.get("details"):
        story.append(P("Details", st["h2"]))
        for para in str(answer["details"]).split("\n\n"):
            if para.strip():
                story.append(P(_cite_to_sup(para.strip()), st["p"]))
    v = answer.get("verification") or {}
    story.append(Spacer(1, 8))
    story.append(P(f"Verification: {v.get('verified', 0)}/{v.get('facts', 0)} facts checked byte-for-byte against the cited passages. "
                   f"Panel verdict: {(answer.get('verdict') or {}).get('verdict', '—')}. Second reader: {_esc(answer.get('second_opinion', ''))}", st["small"]))
    story.append(P(f"Coverage: {_esc(answer.get('coverage', ''))}", st["small"]))
    cited = sorted({int(m) for k in ("headline", "details") for m in re.findall(r"\[#(\d+)\]", str(answer.get(k, "")))}
                   | {s for pt in answer.get("points", []) for s in pt.get("sources", [])})
    srcs = {s["index"]: s for s in answer.get("sources", [])}
    if cited:
        story.append(P("References", st["h2"]))
        for i in cited:
            s = srcs.get(i)
            if s:
                story.append(P(f"<b>{i}.</b> {_esc(s.get('citation'))} — {_esc(s.get('date'))} — {_esc(s.get('placement'))} — sha {_esc((s.get('artifact_sha') or '')[:12])}", st["ref"]))
    doc.build(story)
    return buf.getvalue()


def briefing_pdf(brief: Dict[str, Any], *, user_name: str) -> bytes:
    doc, story, st, buf, P, Spacer, colors = _pdf_doc(f"Morning briefing — {user_name}", f"{brief.get('day')} · generated {str(brief.get('generated_at'))[:16]} UTC")
    story.append(P(_esc(brief.get("headline")), st["h2"]))
    refs: List[str] = []
    for sec in brief.get("sections", []):
        if not sec.get("items"):
            continue
        story.append(P(_esc(sec.get("title")), st["h2"]))
        for it in sec["items"]:
            color = _URG_COLOR.get(it.get("urgency"), "#1f2937")
            ref = ""
            if it.get("source_sha"):
                refs.append(f"{_pl(it.get('property_id'))} — sha {str(it['source_sha'])[:12]}")
                ref = f" <super><font color='#1f6f5f'>{len(refs)}</font></super>"
            prop = f" <font color='#6b6b76'>({_esc(_pl(it['property_id']))})</font>" if it.get("property_id") else ""
            story.append(P(f"<font color='{color}'>■</font> {_esc(it.get('text'))}{prop}{ref}", st["p"]))
    if brief.get("closing"):
        story.append(Spacer(1, 6))
        story.append(P(_esc(brief["closing"]), st["small"]))
    if refs:
        story.append(P("References", st["h2"]))
        for i, r in enumerate(refs, 1):
            story.append(P(f"<b>{i}.</b> {_esc(r)}", st["ref"]))
    doc.build(story)
    return buf.getvalue()
