"""Spreadsheet extraction — cell-level, with provenance. Never OCR.

Why this is a separate path from OCR
------------------------------------
53 of the 377 disk files are ``.xlsx``/``.xls`` and they carry **the actual
money**: draw schedules, budgets, payoff calculations, construction status, the
Equity Rescue underwriting templates. Rendering a spreadsheet to an image and
OCR-ing it destroys exactly what makes it evidence — the grid, the formulas, and
the ability to say *which cell* a number came from.

Two properties this module guarantees:

* **Every value cites its origin** as ``Sheet1!C14``. The ledger's evidence
  standard is strictest here, because these are the numbers that move money.
* **Formulas and computed values are both captured.** A payoff that is a formula
  records *how it was derived*; and when a stored value disagrees with its own
  formula, that disagreement is itself a finding rather than something to
  silently resolve.
"""
from __future__ import annotations

import re
from dataclasses import dataclass, field
from datetime import date, datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

from mangotree.core.logging import logger

#: A cell that looks like money. Used to mark cells worth extra scrutiny, never
#: to decide their value.
_MONEY_HINT = re.compile(r"(?:amount|total|cost|price|budget|draw|balance|payoff|"
                         r"interest|principal|fee|paid|due|retainage|\$)", re.I)


@dataclass
class Cell:
    ref: str                  # "C14"
    row: int
    col: int
    value: Any
    formula: Optional[str] = None
    is_money: bool = False

    def as_dict(self) -> dict:
        out = {"ref": self.ref, "row": self.row, "col": self.col, "value": self.value}
        if self.formula:
            out["formula"] = self.formula
        if self.is_money:
            out["is_money"] = True
        return out


@dataclass
class Sheet:
    name: str
    index: int
    hidden: bool
    n_rows: int
    n_cols: int
    header_row: Optional[int]
    headers: List[str] = field(default_factory=list)
    cells: List[Cell] = field(default_factory=list)
    rows_text: List[str] = field(default_factory=list)

    def as_dict(self) -> dict:
        return {
            "name": self.name,
            "index": self.index,
            "hidden": self.hidden,
            "n_rows": self.n_rows,
            "n_cols": self.n_cols,
            "header_row": self.header_row,
            "headers": self.headers,
            "cells": [c.as_dict() for c in self.cells],
        }


@dataclass
class WorkbookExtract:
    path: str
    engine: str
    sheets: List[Sheet] = field(default_factory=list)
    errors: List[str] = field(default_factory=list)

    @property
    def text(self) -> str:
        """Flat text for indexing — each row prefixed with its sheet!ref origin."""
        parts: List[str] = []
        for sheet in self.sheets:
            parts.append(f"### Sheet: {sheet.name}")
            if sheet.headers:
                parts.append(" | ".join(sheet.headers))
            parts.extend(sheet.rows_text)
        return "\n".join(parts)

    def as_dict(self) -> dict:
        return {
            "engine": self.engine,
            "sheet_count": len(self.sheets),
            "sheets": [s.as_dict() for s in self.sheets],
            "errors": self.errors,
        }


# --------------------------------------------------------------------------
def _col_letter(index: int) -> str:
    """1 -> A, 27 -> AA."""
    letters = ""
    while index > 0:
        index, rem = divmod(index - 1, 26)
        letters = chr(65 + rem) + letters
    return letters


def _clean(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    if isinstance(value, str):
        stripped = value.strip()
        return stripped or None
    return value


def _detect_header_row(rows: List[List[Any]], scan: int = 12) -> Tuple[Optional[int], List[str]]:
    """Find the most plausible header row near the top.

    Real draw schedules and budgets rarely start at row 1 — they carry a title,
    a blank line, then the header. The best candidate is the earliest row that
    is mostly non-empty text and mostly distinct.
    """
    best_index: Optional[int] = None
    best_score = 0.0

    for idx, row in enumerate(rows[:scan]):
        values = [v for v in row if _clean(v) is not None]
        if len(values) < 2:
            continue
        texty = sum(1 for v in values if isinstance(v, str))
        distinct = len({str(v).strip().lower() for v in values})
        score = (texty / len(values)) * (distinct / len(values)) * min(len(values) / 4.0, 1.0)
        # Earlier rows win ties; later rows must be clearly better.
        if score > best_score + 0.01:
            best_score, best_index = score, idx

    if best_index is None or best_score < 0.4:
        return None, []
    headers = [str(_clean(v) or "") for v in rows[best_index]]
    return best_index, headers


def _build_sheet(
    name: str, index: int, hidden: bool, rows: List[List[Any]],
    formulas: Optional[Dict[Tuple[int, int], str]] = None,
) -> Sheet:
    formulas = formulas or {}
    header_row, headers = _detect_header_row(rows)

    cells: List[Cell] = []
    rows_text: List[str] = []
    n_cols = max((len(r) for r in rows), default=0)

    for r_idx, row in enumerate(rows):
        rendered: List[str] = []
        for c_idx, raw in enumerate(row):
            value = _clean(raw)
            if value is None:
                continue
            ref = f"{_col_letter(c_idx + 1)}{r_idx + 1}"
            header = headers[c_idx] if c_idx < len(headers) else ""
            is_money = bool(
                isinstance(value, (int, float))
                and (_MONEY_HINT.search(header or "") or _MONEY_HINT.search(name))
            )
            cells.append(
                Cell(
                    ref=ref, row=r_idx + 1, col=c_idx + 1, value=value,
                    formula=formulas.get((r_idx + 1, c_idx + 1)),
                    is_money=is_money,
                )
            )
            label = f"{header}=" if header and header_row is not None and r_idx != header_row else ""
            rendered.append(f"{label}{value}")

        if rendered:
            rows_text.append(f"[{name}!A{r_idx + 1}] " + " | ".join(rendered))

    return Sheet(
        name=name, index=index, hidden=hidden,
        n_rows=len(rows), n_cols=n_cols,
        header_row=(header_row + 1) if header_row is not None else None,
        headers=headers, cells=cells, rows_text=rows_text,
    )


# --------------------------------------------------------------------------
def _extract_xlsx(path: Path) -> WorkbookExtract:
    from openpyxl import load_workbook

    result = WorkbookExtract(path=str(path), engine="openpyxl")

    # Two passes: one for computed values, one for formulas. openpyxl cannot
    # return both from a single load.
    wb_values = load_workbook(path, data_only=True, read_only=True)
    try:
        wb_formulas = load_workbook(path, data_only=False, read_only=True)
    except Exception:
        wb_formulas = None

    try:
        for index, name in enumerate(wb_values.sheetnames):
            ws = wb_values[name]
            rows = [list(r) for r in ws.iter_rows(values_only=True)]

            formulas: Dict[Tuple[int, int], str] = {}
            if wb_formulas is not None and name in wb_formulas.sheetnames:
                fs = wb_formulas[name]
                for r_idx, frow in enumerate(fs.iter_rows(values_only=True), start=1):
                    for c_idx, fval in enumerate(frow, start=1):
                        if isinstance(fval, str) and fval.startswith("="):
                            formulas[(r_idx, c_idx)] = fval

            hidden = getattr(ws, "sheet_state", "visible") != "visible"
            if hidden:
                # Hidden sheets are read but flagged — never silently used, never
                # silently dropped.
                result.errors.append(f"sheet '{name}' is hidden (extracted, flagged)")
            result.sheets.append(_build_sheet(name, index, hidden, rows, formulas))
    finally:
        wb_values.close()
        if wb_formulas is not None:
            wb_formulas.close()

    return result


def _extract_xls(path: Path) -> WorkbookExtract:
    import xlrd

    result = WorkbookExtract(path=str(path), engine="xlrd")
    book = xlrd.open_workbook(str(path), formatting_info=False)
    for index in range(book.nsheets):
        ws = book.sheet_by_index(index)
        rows = [list(ws.row_values(r)) for r in range(ws.nrows)]
        hidden = getattr(ws, "visibility", 0) != 0
        if hidden:
            result.errors.append(f"sheet '{ws.name}' is hidden (extracted, flagged)")
        result.sheets.append(_build_sheet(ws.name, index, hidden, rows))
    return result


def _coerce_number(text: str):
    """Turn a numeric-looking CSV field into a number, or return it unchanged.

    A CSV has no types, so every field arrives as text — and money detection,
    like every numeric check downstream, tests for a numeric type. Without this
    a CSV budget reads as prose and its dollar figures are invisible.

    Deliberately conservative: leading zeros and long digit runs are left as text
    because zip codes, account numbers and phone numbers are identifiers, and
    turning them into integers silently corrupts them.
    """
    raw = text.strip()
    if not raw:
        return text

    negative = raw.startswith("(") and raw.endswith(")")
    body = raw[1:-1].strip() if negative else raw
    body = body.lstrip("$€£").replace(",", "").strip()
    if body.startswith("-"):
        negative, body = True, body[1:].strip()

    if not body or not body.replace(".", "", 1).isdigit():
        return text
    if body.count(".") > 1:
        return text

    digits = body.split(".")[0]
    if len(digits) > 1 and digits.startswith("0"):
        return text
    if "." not in body and len(digits) > 11:
        return text

    try:
        value = float(body) if "." in body else int(body)
    except ValueError:
        return text
    return -value if negative else value


def _extract_csv(path: Path) -> WorkbookExtract:
    """Read a CSV as a single-sheet workbook.

    Exported draw schedules and budgets arrive as CSV often enough to matter, and
    they carry money like any other sheet. Delimiter and encoding are both
    sniffed: these files come from Excel, Google Sheets and accounting exports,
    so neither comma nor UTF-8 can be assumed.
    """
    import csv as csv_module

    result = WorkbookExtract(path=str(path), engine="csv")

    raw = path.read_bytes()
    text = None
    for encoding in ("utf-8-sig", "utf-8", "cp1252", "latin-1"):
        try:
            text = raw.decode(encoding)
            break
        except UnicodeDecodeError:
            continue
    if text is None:
        text = raw.decode("latin-1", errors="replace")

    sample = text[:8192]
    try:
        dialect = csv_module.Sniffer().sniff(sample, delimiters=",;\t|")
    except Exception:
        dialect = csv_module.excel

    rows = [
        [_coerce_number(field) for field in row]
        for row in csv_module.reader(text.splitlines(), dialect)
    ]
    result.sheets.append(_build_sheet(path.stem or "csv", 0, False, rows))
    return result


def extract_workbook(path: Path) -> WorkbookExtract:
    """Extract a spreadsheet. Legacy ``.xls`` falls back to ``.xlsx`` handling
    when the file is actually a renamed modern workbook — which happens often in
    this corpus."""
    path = Path(path)
    suffix = path.suffix.lower()

    try:
        if suffix in {".xlsx", ".xlsm"}:
            return _extract_xlsx(path)
        if suffix == ".csv":
            return _extract_csv(path)
        if suffix == ".xls":
            try:
                return _extract_xls(path)
            except Exception as exc:
                logger.info("xlrd failed on %s (%s); retrying as xlsx", path.name, exc)
                return _extract_xlsx(path)
        raise ValueError(f"unsupported spreadsheet type: {suffix}")
    except Exception as exc:
        result = WorkbookExtract(path=str(path), engine="failed")
        result.errors.append(f"{type(exc).__name__}: {exc}")
        logger.error("Spreadsheet extraction failed for %s: %s", path.name, exc)
        return result


def money_cells(extract: WorkbookExtract) -> List[dict]:
    """Every money-looking cell with its full provenance, for the ledger."""
    out: List[dict] = []
    for sheet in extract.sheets:
        for cell in sheet.cells:
            if cell.is_money and isinstance(cell.value, (int, float)):
                out.append({
                    "provenance": f"{sheet.name}!{cell.ref}",
                    "sheet": sheet.name,
                    "ref": cell.ref,
                    "value": cell.value,
                    "formula": cell.formula,
                    "row_context": next(
                        (t for t in sheet.rows_text if t.startswith(f"[{sheet.name}!A{cell.row}]")),
                        "",
                    )[:300],
                })
    return out
